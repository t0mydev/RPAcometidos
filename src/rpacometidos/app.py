from flask import Flask, render_template, request, jsonify, send_file
from rpacometidos.lector_excel import procesar_planilla_completa
import json
import io
import openpyxl
import os
from pathlib import Path

# Raíz del proyecto
BASE_DIR = Path(__file__).resolve().parent.parent.parent

# Unifica ambas rutas hacia la carpeta static, ahora mismo inutil ya que flask actúa como api y no como servidor web, pero lo dejamos porque a futuro si servirá los archivos estáticos de Vue
app = Flask(__name__, template_folder="vue", static_folder="vue")

@app.route('/api/procesar-excel', methods=['POST'])
def procesar_excel():
    if 'documento_excel' not in request.files:
        return jsonify({"status": "error", "mensaje": "No se recibió ningún archivo"}), 400

    archivo = request.files['documento_excel']

    if archivo.filename == '':
        return jsonify({"status": "error", "mensaje": "El archivo seleccionado está vacío"}), 400

    
    reporte_validacion = procesar_planilla_completa(archivo)

    return jsonify({
        "status": "completado",
        "resultados": reporte_validacion
    }), 200

@app.route('/api/descargar-excel-corregido', methods=['POST'])
def descargar_excel_corregido():
    if 'documento_excel' not in request.files:
        return jsonify({"status": "error", "mensaje": "No se recibió el archivo original"}), 400
        
    archivo = request.files['documento_excel']
    datos_corregidos_str = request.form.get('reporte_corregido', '[]')
    
    try:
        datos_corregidos = json.loads(datos_corregidos_str)
    except Exception:
        return jsonify({"status": "error", "mensaje": "Formato de datos corregidos inválido"}), 400

    try:
        # Carga el archivo original preservando su estructura
        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active
        
        # Mapea los encabezados para saber las columnas a modificar
        from rpacometidos.lector_excel import (
            buscar_columna,
            HEADER_RUT,
            HEADER_SIGLA,
            HEADER_TIPO_MOVILIZACION,
            HEADER_LUGAR_COMETIDO,
            HEADER_REGION_PRINCIPAL,
            HEADER_REGIONES,
            HEADER_PERSONAL_TRASLADADO,
            HEADER_NOMBRE_APROBADOR,
            HEADER_NOMBRE_FIRMANTES,
            HEADER_TIPO_IMPUTACION_PRESUPUESTARIA,
            HEADER_FALLBACK_CONSIDERANDO,
            HEADER_ATRIBUCION_ACTUAL,
            HEADER_DIAS_SALIDA,
            HEADER_DIAS_100,
            HEADER_DIAS_70,
            HEADER_DIAS_60,
            HEADER_DIAS_50,
            HEADER_DIAS_40,
            HEADER_DIAS_35
        )
        encabezados = {celda.value: celda.column for celda in hoja[2] if celda.value} or {celda.value: celda.column for celda in hoja[1] if celda.value}
        col_rut = buscar_columna(encabezados, HEADER_RUT)
        col_sigla = buscar_columna(encabezados, HEADER_SIGLA)
        col_tipo_movilizacion = buscar_columna(encabezados, HEADER_TIPO_MOVILIZACION)
        col_lugar_cometido = buscar_columna(encabezados, HEADER_LUGAR_COMETIDO)
        col_region_principal = buscar_columna(encabezados, HEADER_REGION_PRINCIPAL)
        col_regiones = buscar_columna(encabezados, HEADER_REGIONES)
        col_personal_trasladado = buscar_columna(encabezados, HEADER_PERSONAL_TRASLADADO)
        col_nombre_aprobador = buscar_columna(encabezados, HEADER_NOMBRE_APROBADOR)
        col_nombre_firmantes = buscar_columna(encabezados, HEADER_NOMBRE_FIRMANTES)
        col_tipo_imputacion_presupuestaria = buscar_columna(encabezados, HEADER_TIPO_IMPUTACION_PRESUPUESTARIA)
        col_fallback_considerando = buscar_columna(encabezados, HEADER_FALLBACK_CONSIDERANDO)
        col_atribucion = buscar_columna(encabezados, HEADER_ATRIBUCION_ACTUAL)
        col_dias_salida = buscar_columna(encabezados, HEADER_DIAS_SALIDA)
        col_dias_100 = buscar_columna(encabezados, HEADER_DIAS_100)
        col_dias_70 = buscar_columna(encabezados, HEADER_DIAS_70)
        col_dias_60 = buscar_columna(encabezados, HEADER_DIAS_60)
        col_dias_50 = buscar_columna(encabezados, HEADER_DIAS_50)
        col_dias_40 = buscar_columna(encabezados, HEADER_DIAS_40)
        col_dias_35 = buscar_columna(encabezados, HEADER_DIAS_35)

        # Modifica los valores
        for registro in datos_corregidos:
            fila_indice = registro.get("numero_fila_excel")
            if not fila_indice:
                continue
            
            if col_rut and "rut" in registro:
                hoja.cell(row=fila_indice, column=col_rut).value = registro["rut"]
            if col_sigla and "sigla" in registro:
                hoja.cell(row=fila_indice, column=col_sigla).value = registro["sigla"]
            if col_tipo_movilizacion and "tipo_movilizacion" in registro:
                hoja.cell(row=fila_indice, column=col_tipo_movilizacion).value = registro["tipo_movilizacion"]
            if col_lugar_cometido and "lugar_cometido" in registro:
                hoja.cell(row=fila_indice, column=col_lugar_cometido).value = registro["lugar_cometido"]
            if col_region_principal and "region_principal" in registro:
                hoja.cell(row=fila_indice, column=col_region_principal).value = registro["region_principal"]
            if col_regiones and "regiones" in registro:
                hoja.cell(row=fila_indice, column=col_regiones).value = registro["regiones"]
            if col_personal_trasladado and "personal_trasladado" in registro:
                hoja.cell(row=fila_indice, column=col_personal_trasladado).value = registro["personal_trasladado"]
            if col_nombre_aprobador and "nombre_aprobador" in registro:
                hoja.cell(row=fila_indice, column=col_nombre_aprobador).value = registro["nombre_aprobador"]
            if col_nombre_firmantes and "nombre_firmantes" in registro:
                hoja.cell(row=fila_indice, column=col_nombre_firmantes).value = registro["nombre_firmantes"]
            if col_tipo_imputacion_presupuestaria and "tipo_imputacion_presupuestaria" in registro:
                hoja.cell(row=fila_indice, column=col_tipo_imputacion_presupuestaria).value = registro["tipo_imputacion_presupuestaria"]
            if col_fallback_considerando and "fallback_considerando" in registro:
                hoja.cell(row=fila_indice, column=col_fallback_considerando).value = registro["fallback_considerando"]
            if col_atribucion and "atribucion" in registro:
                hoja.cell(row=fila_indice, column=col_atribucion).value = registro["atribucion"]
            if col_dias_salida and "dias_salida" in registro:
                hoja.cell(row=fila_indice, column=col_dias_salida).value = registro["dias_salida"]
            if col_dias_100 and "dias_100" in registro:
                hoja.cell(row=fila_indice, column=col_dias_100).value = registro["dias_100"]
            if col_dias_70 and "dias_70" in registro:
                hoja.cell(row=fila_indice, column=col_dias_70).value = registro["dias_70"]
            if col_dias_60 and "dias_60" in registro:
                hoja.cell(row=fila_indice, column=col_dias_60).value = registro["dias_60"]
            if col_dias_50 and "dias_50" in registro:
                hoja.cell(row=fila_indice, column=col_dias_50).value = registro["dias_50"]
            if col_dias_40 and "dias_40" in registro:
                hoja.cell(row=fila_indice, column=col_dias_40).value = registro["dias_40"]
            if col_dias_35 and "dias_35" in registro:
                hoja.cell(row=fila_indice, column=col_dias_35).value = registro["dias_35"]

        # Guarda el archivo corregido en un buffer en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        nombre_original = archivo.filename or "planilla.xlsx"
        nombre_corregido = nombre_original.rsplit('.', 1)[0] + "_corregido.xlsx"
        
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=nombre_corregido
        )
    except Exception as e:
        return jsonify({"status": "error", "mensaje": f"Error al generar el archivo: {str(e)}"}), 500
    
#Codigo que quedó de la simulación de automatización que se hizo para la presentación, puede que sirva en el futuro para la automatización real, por ahora no se usa
import subprocess
import sys

@app.route('/api/empezar-automatizacion', methods=['POST'])
def empezar_automatizacion():
    try:
        path_progreso = BASE_DIR / "progreso_automatizacion.json"
        path_datos = BASE_DIR / "datos_automatizacion.json"

        # Borra el progreso anterior si existe
        if path_progreso.exists():
            try:
                path_progreso.unlink()
            except Exception:
                pass

        # Recupera los registros corregidos desde la vista
        datos = request.json or []
        
        # Le damos formato a los datos para el robot de Playwright
        datos_robot = []
        for registro in datos:
            rut = registro.get("rut")
            sigla = registro.get("sigla")
            fechainicio = registro.get("fechainicio")
            fechatermino = registro.get("fechatermino")
            tipo_movilizacion = registro.get("tipo_movilizacion")
            personal_trasladado = registro.get("personal_trasladado")
            fallback_considerando = registro.get("fallback_considerando")
            lugar_cometido = registro.get("lugar_cometido")
            regiones = registro.get("regiones")
            atribucion = registro.get("atribucion")
            dias_salida = registro.get("dias_salida")
            dias_100 = registro.get("dias_100")
            dias_70 = registro.get("dias_70")
            dias_60 = registro.get("dias_60")
            dias_50 = registro.get("dias_50")
            dias_40 = registro.get("dias_40")
            dias_35 = registro.get("dias_35")
            tipo_imputacion_presupuestaria = registro.get("tipo_imputacion_presupuestaria")
            
            # Por defecto aprobado, a menos que falten datos esenciales
            estado = "aprobado"
            if not rut or not sigla:
                estado = "pendiente"
                
            datos_robot.append({
                "rut": rut or "",
                "sigla": sigla or "",
                "fechainicio": fechainicio or "",
                "fechatermino": fechatermino or "",
                "tipo_movilizacion": tipo_movilizacion or "",
                "personal_trasladado": personal_trasladado or "",
                "fallback_considerando": fallback_considerando or "",
                "lugar_cometido": lugar_cometido or "",
                "regiones": regiones or "",
                "atribucion": atribucion or "",
                "dias_salida": dias_salida or "",
                "dias_100": dias_100 if dias_100 is not None else "",
                "dias_70": dias_70 if dias_70 is not None else "",
                "dias_60": dias_60 if dias_60 is not None else "",
                "dias_50": dias_50 if dias_50 is not None else "",
                "dias_40": dias_40 if dias_40 is not None else "",
                "dias_35": dias_35 if dias_35 is not None else "",
                "tipo_imputacion_presupuestaria": tipo_imputacion_presupuestaria or "",
                "accion": estado
            })
            
        # Guarda los registros en datos_automatizacion.json
        with open(path_datos, "w", encoding="utf-8") as f:
            json.dump(datos_robot, f, ensure_ascii=False, indent=4)
            
        # Ejecutam el orquestador de robots en segundo plano
        subprocess.Popen([sys.executable, "-m", "rpacometidos.robots.orquestador"], cwd=str(BASE_DIR))
        
        return jsonify({"status": "iniciado", "mensaje": "La automatización se ha iniciado correctamente."}), 200
        
    except Exception as e:
        return jsonify({"status": "error", "mensaje": f"Error al iniciar el robot: {str(e)}"}), 500

@app.route('/api/progreso-automatizacion', methods=['GET'])
def progreso_automatizacion():
    path_progreso = BASE_DIR / "progreso_automatizacion.json"
    if path_progreso.exists():
        try:
            with open(path_progreso, "r", encoding="utf-8") as f:
                datos = json.load(f)
            return jsonify(datos), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        return jsonify({"estado": "no_iniciado"}), 200


if __name__ == '__main__':
    app.run(debug=True, port=5000)
