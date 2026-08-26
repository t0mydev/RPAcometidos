import openpyxl
import re
from rpacometidos.procesador_datos import validar_registro

def normalizar_texto(texto):
    if not texto:
        return ""
    t = str(texto).lower().strip()
    t = t.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    return re.sub(r'\s+', ' ', t)

# Nombres de encabezados exactos de la plantilla oficial
HEADER_RUT = "RUT Funcionario (Chofer)"
HEADER_FECHA_INICIO = "Fecha inicio (DD/MM/YYYY)"
HEADER_FECHA_TERMINO = "Fecha Termino (DD/MM/YYYY)"
HEADER_TIPO_MOVILIZACION = "Tipo de movilizacion"
HEADER_SIGLA = "Sigla Vehiculo"
HEADER_LUGAR_COMETIDO = "Lugar Cometido"
HEADER_REGION_PRINCIPAL = "Región Principal"
HEADER_REGIONES = "Region/es"
HEADER_PERSONAL_TRASLADADO = "Personal Trasladado"
HEADER_NOMBRE_APROBADOR = "Nombre Aprobador"
HEADER_NOMBRE_FIRMANTES = "Nombre Firmante/s"
HEADER_TIPO_IMPUTACION_PRESUPUESTARIA = "Tipo Imputacion Presupuestaria"
HEADER_FALLBACK_CONSIDERANDO = "Fallback Considerando"
HEADER_ATRIBUCION_ACTUAL = "Atribucion"
HEADER_DIAS_SALIDA = "Dias de salida"
HEADER_DIAS_100 = "Dias al 100%"
HEADER_DIAS_70 = "Dias al 70%"
HEADER_DIAS_60 = "Dias al 60%"
HEADER_DIAS_50 = "Dias al 50%"
HEADER_DIAS_40 = "Dias al 40%"
HEADER_DIAS_35 = "Dias al 35%"

def buscar_columna(encabezados, nombre_encabezado):
    norm_target = normalizar_texto(nombre_encabezado)
    for celda_val, col_num in encabezados.items():
        if not celda_val:
            continue
        celda_norm = normalizar_texto(celda_val)
        if celda_norm == norm_target or norm_target in celda_norm or celda_norm in norm_target:
            return col_num
    return None

def procesar_planilla_completa(archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    hoja = wb.active
    
    reporte_final = []

    # Busca la fila de encabezados probando Fila 2 primero (plantilla oficial) y Fila 1 como fallback
    encabezados = {}
    fila_encabezados = 2
    for r_idx in [2, 1]:
        posibles = {celda.value: celda.column for celda in hoja[r_idx] if celda.value}
        if any("rut" in normalizar_texto(v) for v in posibles.keys()):
            encabezados = posibles
            fila_encabezados = r_idx
            break
    if not encabezados:
        encabezados = {celda.value: celda.column for celda in hoja[1] if celda.value}
        fila_encabezados = 1

    col_rut = buscar_columna(encabezados, HEADER_RUT)
    col_fechainicio = buscar_columna(encabezados, HEADER_FECHA_INICIO)
    col_fechatermino = buscar_columna(encabezados, HEADER_FECHA_TERMINO)
    col_tipo_movilizacion = buscar_columna(encabezados, HEADER_TIPO_MOVILIZACION)
    col_sigla = buscar_columna(encabezados, HEADER_SIGLA)
    col_lugar_cometido = buscar_columna(encabezados, HEADER_LUGAR_COMETIDO)
    col_region_principal = buscar_columna(encabezados, HEADER_REGION_PRINCIPAL)
    col_regiones = buscar_columna(encabezados, HEADER_REGIONES)
    col_personal_trasladado = buscar_columna(encabezados, HEADER_PERSONAL_TRASLADADO)
    col_nombre_aprobador = buscar_columna(encabezados, HEADER_NOMBRE_APROBADOR)
    col_nombre_firmantes = buscar_columna(encabezados, HEADER_NOMBRE_FIRMANTES)
    col_tipo_imputacion_presupuestaria = buscar_columna(encabezados, HEADER_TIPO_IMPUTACION_PRESUPUESTARIA)
    col_fallback_considerando = buscar_columna(encabezados, HEADER_FALLBACK_CONSIDERANDO)
    col_atribucion = buscar_columna(encabezados, HEADER_ATRIBUCION_ACTUAL)
    col_dias_salida = buscar_columna(encabezados, HEADER_DIAS_SALIDA)
    col_dias_100 = buscar_columna(encabezados, HEADER_DIAS_100)
    col_dias_70 = buscar_columna(encabezados, HEADER_DIAS_70)
    col_dias_60 = buscar_columna(encabezados, HEADER_DIAS_60)
    col_dias_50 = buscar_columna(encabezados, HEADER_DIAS_50)
    col_dias_40 = buscar_columna(encabezados, HEADER_DIAS_40)
    col_dias_35 = buscar_columna(encabezados, HEADER_DIAS_35)

    if not col_rut and not col_sigla:
        raise ValueError("No se encontraron las columnas principales en la plantilla Excel cargada.")

    # Itera desde la fila siguiente a la de encabezados hasta la última fila con datos
    for fila_indice in range(fila_encabezados + 1, hoja.max_row + 1):
        rut = hoja.cell(row=fila_indice, column=col_rut).value if col_rut else None
        sigla = hoja.cell(row=fila_indice, column=col_sigla).value if col_sigla else None
        fechainicio = hoja.cell(row=fila_indice, column=col_fechainicio).value if col_fechainicio else None
        fechatermino = hoja.cell(row=fila_indice, column=col_fechatermino).value if col_fechatermino else None
        tipo_movilizacion = hoja.cell(row=fila_indice, column=col_tipo_movilizacion).value if col_tipo_movilizacion else None
        lugar_cometido = hoja.cell(row=fila_indice, column=col_lugar_cometido).value if col_lugar_cometido else None
        region_principal = hoja.cell(row=fila_indice, column=col_region_principal).value if col_region_principal else None
        regiones = hoja.cell(row=fila_indice, column=col_regiones).value if col_regiones else None
        personal_trasladado = hoja.cell(row=fila_indice, column=col_personal_trasladado).value if col_personal_trasladado else None
        nombre_aprobador = hoja.cell(row=fila_indice, column=col_nombre_aprobador).value if col_nombre_aprobador else None
        nombre_firmantes = hoja.cell(row=fila_indice, column=col_nombre_firmantes).value if col_nombre_firmantes else None
        tipo_imputacion_presupuestaria = hoja.cell(row=fila_indice, column=col_tipo_imputacion_presupuestaria).value if col_tipo_imputacion_presupuestaria else None
        fallback_considerando = hoja.cell(row=fila_indice, column=col_fallback_considerando).value if col_fallback_considerando else None
        atribucion = hoja.cell(row=fila_indice, column=col_atribucion).value if col_atribucion else None
        dias_salida = hoja.cell(row=fila_indice, column=col_dias_salida).value if col_dias_salida else None
        dias_100 = hoja.cell(row=fila_indice, column=col_dias_100).value if col_dias_100 else None
        dias_70 = hoja.cell(row=fila_indice, column=col_dias_70).value if col_dias_70 else None
        dias_60 = hoja.cell(row=fila_indice, column=col_dias_60).value if col_dias_60 else None
        dias_50 = hoja.cell(row=fila_indice, column=col_dias_50).value if col_dias_50 else None
        dias_40 = hoja.cell(row=fila_indice, column=col_dias_40).value if col_dias_40 else None
        dias_35 = hoja.cell(row=fila_indice, column=col_dias_35).value if col_dias_35 else None

        # Ignora fila totalmente en blanco
        if (rut is None and sigla is None and fechainicio is None and fechatermino is None and 
            tipo_movilizacion is None and lugar_cometido is None and region_principal is None and 
            regiones is None and personal_trasladado is None and nombre_aprobador is None and 
            nombre_firmantes is None and tipo_imputacion_presupuestaria is None and 
            fallback_considerando is None and atribucion is None and dias_salida is None and
            dias_100 is None and dias_70 is None and dias_60 is None and dias_50 is None and 
            dias_40 is None and dias_35 is None):
            continue
        
        datos_fila = {
            "rut": rut,
            "sigla": sigla,
        }
        
        resultado_fila = validar_registro(datos_fila)
        
        resultado_fila["rut"] = rut
        resultado_fila["sigla"] = sigla
        resultado_fila["fechainicio"] = fechainicio
        resultado_fila["fechatermino"] = fechatermino
        resultado_fila["tipo_movilizacion"] = tipo_movilizacion
        resultado_fila["lugar_cometido"] = lugar_cometido
        resultado_fila["region_principal"] = region_principal
        resultado_fila["regiones"] = regiones
        resultado_fila["personal_trasladado"] = personal_trasladado
        resultado_fila["nombre_aprobador"] = nombre_aprobador
        resultado_fila["nombre_firmantes"] = nombre_firmantes
        resultado_fila["tipo_imputacion_presupuestaria"] = tipo_imputacion_presupuestaria
        resultado_fila["fallback_considerando"] = fallback_considerando
        resultado_fila["atribucion"] = atribucion
        resultado_fila["dias_salida"] = dias_salida
        resultado_fila["dias_100"] = dias_100
        resultado_fila["dias_70"] = dias_70
        resultado_fila["dias_60"] = dias_60
        resultado_fila["dias_50"] = dias_50
        resultado_fila["dias_40"] = dias_40
        resultado_fila["dias_35"] = dias_35
        resultado_fila["numero_fila_excel"] = fila_indice

        reporte_final.append(resultado_fila)

    return reporte_final